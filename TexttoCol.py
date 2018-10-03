import openpyxl as xl
import re as r

fl = xl.load_workbook('example.xlsx')
ws = fl['Sheet1']

mylist = []

for row in ws.iter_rows('A{}:A{}'.format(ws.min_row, ws.max_row)):
    for cell in row:
        regex = r.compile(r'(?:(?!\s\d).)*')
        seprate = regex.search(cell.value)
        print(seprate.group())
                
